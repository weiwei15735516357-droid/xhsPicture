from datetime import datetime
import csv
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFilter, ImageFont

from backend.services.asset_registry import AssetRegistry


XHS_WIDTH = 1080
XHS_HEIGHT = 1440


class PerspectiveComposer:
    def compose_batch(
        self,
        project_dir: Path,
        scene_path: Path,
        overlay_paths: list[Path],
        points: list[dict[str, float]],
        opacity: float = 1.0,
        shadow: bool = True,
        progress_callback: Callable[[int, int, str], None] | None = None,
    ) -> dict[str, Any]:
        if len(points) != 4:
            raise ValueError("透视合成需要 4 个角点")
        if not overlay_paths:
            raise ValueError("请至少选择一张叠图")
        output_dir = project_dir / "compositions" / datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir.mkdir(parents=True, exist_ok=True)
        registry = AssetRegistry(project_dir)
        scene = self._cover_to_canvas(Image.open(scene_path).convert("RGB"))
        quad = self._normalized_quad(points)
        assets = []
        total = len(overlay_paths)
        for index, overlay_path in enumerate(overlay_paths, start=1):
            overlay = Image.open(overlay_path).convert("RGBA")
            composed = self._compose_one(scene, overlay, quad, opacity=opacity, shadow=shadow)
            output_path = output_dir / f"{overlay_path.stem}_透视合成_{index:03d}.png"
            composed.save(output_path)
            assets.append(registry.add_asset(output_path, "透视合成图", str(overlay_path)))
            if progress_callback:
                progress_callback(index, total, f"正在生成透视合成 {index}/{total}")
        return {"assets": assets}

    def compose_text_batch(
        self,
        project_dir: Path,
        scene_path: Path,
        excel_path: Path,
        text_options: dict[str, Any] | None = None,
        text_rows: list[dict[str, Any]] | None = None,
        progress_callback: Callable[[int, int, str], None] | None = None,
    ) -> dict[str, Any]:
        rows = text_rows or self.read_excel_rows(excel_path)
        if not rows:
            raise ValueError("Excel 中没有可用数据")
        output_dir = project_dir / "compositions" / datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir.mkdir(parents=True, exist_ok=True)
        registry = AssetRegistry(project_dir)
        scene = self._cover_to_canvas(Image.open(scene_path).convert("RGB"))
        assets = []
        total = len(rows)
        for index, row in enumerate(rows, start=1):
            row_options = row.get("text_options") or text_options or {}
            composed = self._compose_text(scene, str(row["title"]), row_options)
            safe_id = self._safe_filename(str(row["product_id"]))
            output_path = output_dir / f"{safe_id}.png"
            output_path = self._unique_path(output_path)
            composed.save(output_path)
            assets.append(registry.add_asset(output_path, "透视文字图", str(excel_path)))
            if progress_callback:
                progress_callback(index, total, f"正在生成文字叠图 {index}/{total}")
        return {"assets": assets}

    def _compose_one(
        self,
        scene: Image.Image,
        overlay: Image.Image,
        quad: list[tuple[int, int]],
        opacity: float,
        shadow: bool,
    ) -> Image.Image:
        canvas = scene.convert("RGBA")
        if shadow:
            shadow_layer = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
            shadow_draw = ImageDraw.Draw(shadow_layer)
            shadow_draw.polygon([(x + 10, y + 14) for x, y in quad], fill=(0, 0, 0, 72))
            canvas.alpha_composite(shadow_layer.filter(ImageFilter.GaussianBlur(14)))
        warped = self._warp_overlay(overlay, quad)
        if opacity < 1:
            alpha = warped.getchannel("A").point(lambda value: int(value * opacity))
            warped.putalpha(alpha)
        canvas.alpha_composite(warped)
        return canvas.convert("RGB")

    def _compose_text(self, scene: Image.Image, title: str, options: dict[str, Any]) -> Image.Image:
        canvas = scene.convert("RGBA")
        overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)
        x = int(options.get("x", 118))
        y = int(options.get("y", 386))
        font_size = int(options.get("font_size", 92))
        stroke_width = int(options.get("stroke_width", 0))
        font = self._title_font(
            font_size=font_size,
            bold=bool(options.get("bold", True)),
            font_family=str(options.get("font_family", "msyh")),
        )
        lines = self._wrap_text(title, font, max(80, XHS_WIDTH - x - 54))
        line_height = int(font_size * 1.25)
        fill = self._hex_to_rgba(str(options.get("color", "#000000")))
        stroke_fill = self._hex_to_rgba(str(options.get("stroke_color", "#ffffff")))
        for line in lines[:5]:
            draw.text(
                (x, y),
                line,
                fill=fill,
                font=font,
                stroke_width=stroke_width,
                stroke_fill=stroke_fill,
            )
            y += line_height
        canvas.alpha_composite(overlay)
        return canvas.convert("RGB")

    def _warp_overlay(self, overlay: Image.Image, quad: list[tuple[int, int]]) -> Image.Image:
        coeffs = self._perspective_coefficients(
            quad,
            [(0, 0), (overlay.width, 0), (overlay.width, overlay.height), (0, overlay.height)],
        )
        return overlay.transform(
            (XHS_WIDTH, XHS_HEIGHT),
            Image.Transform.PERSPECTIVE,
            coeffs,
            Image.Resampling.BICUBIC,
        )

    def _cover_to_canvas(self, image: Image.Image) -> Image.Image:
        canvas_ratio = XHS_WIDTH / XHS_HEIGHT
        image_ratio = image.width / image.height
        if image_ratio > canvas_ratio:
            new_height = XHS_HEIGHT
            new_width = int(new_height * image_ratio)
        else:
            new_width = XHS_WIDTH
            new_height = int(new_width / image_ratio)
        resized = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
        left = (new_width - XHS_WIDTH) // 2
        top = (new_height - XHS_HEIGHT) // 2
        return resized.crop((left, top, left + XHS_WIDTH, top + XHS_HEIGHT))

    def read_excel_rows(self, excel_path: Path) -> list[dict[str, str]]:
        if excel_path.suffix.lower() == ".csv":
            with excel_path.open("r", encoding="utf-8-sig", newline="") as file:
                return self._normalize_rows(list(csv.DictReader(file)))
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records = []
        for row in rows[1:]:
            records.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
        return self._normalize_rows(records)

    def _normalize_rows(self, records: list[dict[str, Any]]) -> list[dict[str, str]]:
        normalized = []
        for record in records:
            product_id = self._pick_field(record, ["商品id", "商品ID", "product_id", "Product ID", "id", "ID"])
            title = self._pick_field(record, ["标题", "title", "Title", "商品标题"])
            if product_id and title:
                normalized.append({"product_id": product_id, "title": title})
        return normalized

    def _pick_field(self, record: dict[str, Any], names: list[str]) -> str:
        lookup = {str(key).strip().lower(): value for key, value in record.items()}
        for name in names:
            value = lookup.get(name.lower())
            if value is not None and str(value).strip():
                return str(value).strip()
        return ""

    def _title_font(
        self,
        font_size: int = 56,
        bold: bool = True,
        font_family: str = "msyh",
    ) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        font_map = {
            "msyh": [Path("C:/Windows/Fonts/msyhbd.ttc") if bold else Path("C:/Windows/Fonts/msyh.ttc")],
            "simhei": [Path("C:/Windows/Fonts/simhei.ttf")],
            "simsun": [Path("C:/Windows/Fonts/simsun.ttc")],
            "kaiti": [Path("C:/Windows/Fonts/simkai.ttf")],
            "dengxian": [Path("C:/Windows/Fonts/Dengb.ttf") if bold else Path("C:/Windows/Fonts/Deng.ttf")],
        }
        candidates = []
        candidates.extend(font_map.get(font_family, []))
        candidates.extend([
            Path("C:/Windows/Fonts/msyhbd.ttc") if bold else Path("C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/msyh.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
        ])
        for path in candidates:
            if path.exists():
                return ImageFont.truetype(str(path), font_size)
        return ImageFont.load_default()

    def _hex_to_rgba(self, value: str) -> tuple[int, int, int, int]:
        text = value.strip().lstrip("#")
        if len(text) == 3:
            text = "".join(char * 2 for char in text)
        if len(text) != 6:
            return (0, 0, 0, 255)
        try:
            return (int(text[0:2], 16), int(text[2:4], 16), int(text[4:6], 16), 255)
        except ValueError:
            return (0, 0, 0, 255)

    def _wrap_text(self, text: str, font: ImageFont.FreeTypeFont | ImageFont.ImageFont, max_width: int) -> list[str]:
        lines = []
        current = ""
        measure = ImageDraw.Draw(Image.new("RGB", (1, 1)))
        for char in text:
            candidate = current + char
            bbox = measure.textbbox((0, 0), candidate, font=font)
            if current and bbox[2] - bbox[0] > max_width:
                lines.append(current)
                current = char
            else:
                current = candidate
        if current:
            lines.append(current)
        return lines or [""]

    def _safe_filename(self, value: str) -> str:
        safe = "".join(char if char not in r'\/:*?"<>|' else "_" for char in value).strip()
        return safe or "未命名"

    def _unique_path(self, path: Path) -> Path:
        if not path.exists():
            return path
        for index in range(1, 10000):
            candidate = path.with_name(f"{path.stem}_{index:03d}{path.suffix}")
            if not candidate.exists():
                return candidate
        raise ValueError("输出文件重名过多，请更换输出目录")

    def _normalized_quad(self, points: list[dict[str, float]]) -> list[tuple[int, int]]:
        quad = []
        for point in points:
            x = max(0, min(1, float(point["x"])))
            y = max(0, min(1, float(point["y"])))
            quad.append((int(x * XHS_WIDTH), int(y * XHS_HEIGHT)))
        return quad

    def _perspective_coefficients(
        self,
        source: list[tuple[int, int]],
        target: list[tuple[int, int]],
    ) -> tuple[float, float, float, float, float, float, float, float]:
        matrix = []
        vector = []
        for (x, y), (u, v) in zip(source, target):
            matrix.append([x, y, 1, 0, 0, 0, -u * x, -u * y])
            vector.append(u)
            matrix.append([0, 0, 0, x, y, 1, -v * x, -v * y])
            vector.append(v)
        return tuple(self._solve_linear_system(matrix, vector))

    def _solve_linear_system(self, matrix: list[list[float]], vector: list[float]) -> list[float]:
        rows = [row[:] + [value] for row, value in zip(matrix, vector)]
        size = len(vector)
        for pivot_index in range(size):
            pivot_row = max(range(pivot_index, size), key=lambda row: abs(rows[row][pivot_index]))
            if abs(rows[pivot_row][pivot_index]) < 1e-9:
                raise ValueError("透视角点无效，请调整四个角点")
            rows[pivot_index], rows[pivot_row] = rows[pivot_row], rows[pivot_index]
            pivot = rows[pivot_index][pivot_index]
            rows[pivot_index] = [value / pivot for value in rows[pivot_index]]
            for row_index in range(size):
                if row_index == pivot_index:
                    continue
                factor = rows[row_index][pivot_index]
                rows[row_index] = [
                    value - factor * pivot_value
                    for value, pivot_value in zip(rows[row_index], rows[pivot_index])
                ]
        return [row[-1] for row in rows]
